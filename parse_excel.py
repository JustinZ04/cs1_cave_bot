#!/usr/bin/env python
# Written by Matthew Villegas & Justin Zabel
# Finds the current day and time and uses that data to search through a spreadsheet containing the office hours of
# TA's for the class. Returns a list of all TA's who are currently holding office hours.

import time
import calendar
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date


# Cls will either be 1 or 2. 1 for CS1, 2 for CS2.
def parse(cls):
    if cls == 1:
        # Need to change tuple bounds in this dictionary to account for the new spreadsheet.
        weekday_dict = {'Monday': ('C', 'M'), 'Tuesday': ('Q', 'W'), 'Wednesday': ('AA', 'AG'),
                        'Thursday': ('AK', 'AO'), 'Friday': ('AU', 'AY')}
        start_times = {'Monday': 930, 'Tuesday': 1400, 'Wednesday': 930, 'Thursday': 1400,
                       'Friday': 1100}
        end_times = {'Monday': 1900, 'Tuesday': 1700, 'Wednesday': 1700, 'Thursday': 1700,
                     'Friday': 1600}
        wb = load_workbook(filename='cs1_office_hours.xlsx', data_only=True, read_only=True)
    else:
        # Need to change tuple bounds in this dictionary to account for the new spreadsheet.
        weekday_dict = {'Monday': ('C', 'D'), 'Tuesday': ('G', 'M'), 'Wednesday': ('Q', 'S'), 'Thursday': ('Y', 'AE'),
                        'Friday': ('AJ', 'AK')}
        start_times = {'Monday': 1600, 'Tuesday': 1300, 'Wednesday': 1200, 'Thursday': 1300,
                       'Friday': 1000}
        end_times = {'Monday': 1930, 'Tuesday': 1630, 'Wednesday': 1930, 'Thursday': 1630,
                     'Friday': 1000}
        wb = load_workbook(filename='cs2_office_hours.xlsx', data_only=True, read_only=True)

    ws = wb['Office Hours']  # Will have to change the name of the worksheet each time.
    ta_list = []

    # Get the current weekday
    day = date.today()
    weekday = calendar.day_name[day.weekday()]

    # Special case for CS2 Friday because there are no office hours held.
    if weekday == 'Friday' and cls == 2:
        return None

    if weekday == 'Saturday' or weekday == 'Sunday':
        return None

    # Get the current time into 24 hour integer format.
    cur_time = time.localtime()
    cur_time = time.strftime("%H%M", cur_time)
    cur_time = int(cur_time)

    # Range of times on the spreadsheet. This will need to be looked at for each spreadsheet.
    if cur_time < start_times[weekday] or cur_time > end_times[weekday]:
        return None

    time_range = None

    # Set the start and end rows based on which spreadsheet is being used. Account for exclusivity of the
    # range function on the upper bound. Also set the time of the earliest TA's office hours to use later
    # when converting the times to 24 hour format.
    if cls == 1:
        start_time = 930
        start_row = 5
        end_row = 52

    else:
        start_time = 1000
        start_row = 5
        end_row = 46

    for j in range(start_row, end_row):
        cell = str(get_column_letter(1) + str(j))

        # Format the time stored in the cell to just a number.
        if str(ws[cell].value) != "None":
            s = str(ws[cell].value)
            s = s.split(' - ')
            s[0] = s[0].replace(":", "")
            s[1] = s[1].replace(":", "")
            s[1] = s[1].replace(" AM", "")
            s[1] = s[1].replace(" PM", "")

            # Converts cur_time from 24 hour time to easily compare
            # against values in the time column of the spreadsheet.
            if cur_time > 1259:
                temp = cur_time - 1200
            else:
                temp = cur_time
            if abs(temp - int(s[0])) < 30:
                time_range = ws[cell].row
                break

    if time_range is None:
        return None

    found_ta = False

    # Search through the spreadsheet from bottom to top, looking for the first TA for which the current time falls
    # between their beginning and ending time. Because of the format of the spreadsheet, as soon as 1 TA is found
    # in a column we can move 2 columns to the right.
    i = column_index_from_string((weekday_dict[weekday])[0])
    while i <= column_index_from_string((weekday_dict[weekday])[1]):
        j = time_range

        while j >= start_row:  # Will need to change to account for lower bound of each spreadsheet.
            col = get_column_letter(i)
            row = str(j)
            cell = "".join((col, row))
            s = ws[cell].value

            if s is None:
                j -= 1
                continue

            office_hour = re.findall(r'\d?\d:\d{2} - \d?\d:\d{2}', s)

            if office_hour is not None:
                if len(office_hour) > 0:
                    # More character replacement to get each TA's beginning and ending office hour time into
                    # a number only format.
                    office_hour = office_hour[0].split(' - ')

                    office_hour[0] = int(office_hour[0].replace(":", ""))
                    office_hour[1] = int(office_hour[1].replace(":", ""))
                    if office_hour[0] < start_time:  # Will need to change for new lower TA hour bound.
                        office_hour[0] = office_hour[0] + 1200  # Add 1200 to get into 24 hour format.

                    if office_hour[1] < start_time:  # Will need to change for new lower TA hour bound.
                        office_hour[1] = office_hour[1] + 1200  # Add 1200 to get into 24 hour format.

                    if office_hour[0] <= cur_time < office_hour[1]:
                        # Find a TA and add them to a list of current TA's.
                        ta_list.append(ws[cell].value)
                        found_ta = True
                        i += 1
                        break
                    else:
                        i += 1
                        break
            j = j - 1
        i = i + 1

    if not found_ta:
        return None

    # Return the list for use in the discord_bot script.
    return ta_list


# Check cs1 hours if testing.
if __name__ == '__main__':
    print(parse(1))
